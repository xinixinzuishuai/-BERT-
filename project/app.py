# ==============================================================================
# 文件: app.py
# ==============================================================================
# 本文件是整个项目的【后端核心】，基于 Flask 框架提供 RESTful API 和页面路由。
# 是前端页面、爬虫模块、BERT 模型之间的"调度中心"。
#
# 【功能索引】（Ctrl+F 搜索关键词可快速定位）
#   [路由-页面]   / (首页), /detail/<topic> (详情页)
#   [路由-预测]   /predict (单条文本情感预测)
#   [路由-分析]   /analyze (批量分析评论)
#   [路由-统计]   /api/sentiment_stats/<topic> (情感统计-饼图+折线图数据)
#   [路由-词云]   /api/wordcloud/<topic> (词云数据)
#   [路由-爬虫]   /api/run_spider (启动爬虫-SSE流式响应)
#   [路由-状态]   /api/spider_status (爬虫运行状态)
#   [路由-话题]   /api/hot_topics (获取已有话题列表)
#   [路由-导出]   /export/<topic>/<format> (导出CSV/Excel)
#   [路由-摘要]   /export_summary/<topic> (导出统计摘要)
#   [路由-健康]   /health (健康检查)
#   [SSE]         Server-Sent Events 流式进度推送
#   [分词]        jieba TextRank 关键词提取
#   [反爬]        CORS 跨域配置
# ==============================================================================

from flask import Flask, request, jsonify, render_template, send_file, make_response, Response, stream_with_context
from flask_cors import CORS
from model_inference import analyzer
from config import Config
from models import db, Comment
from datetime import datetime
from sqlalchemy import func
from urllib.parse import quote
import re
import os
import pandas as pd
import io
import threading
import queue
import jieba.analyse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Flask 应用初始化
# --------------------------------------------------------------------------
app = Flask(__name__)
app.config.from_object(Config)

# [反爬] CORS: 跨域资源共享配置
# 答辩要点：前后端分离架构下，浏览器有"同源策略"限制，
# 即前端页面（如 localhost:8000）不能请求其他域名的 API。
# CORS(app) 允许所有域名访问本服务器的 API，解决跨域问题。
CORS(app)

# 将 SQLAlchemy 绑定到 Flask 应用
db.init_app(app)

# --------------------------------------------------------------------------
# 全局爬虫状态字典
# --------------------------------------------------------------------------
# 答辩要点：使用全局字典而非数据库来存储爬虫状态，因为状态是"临时性"的——
# 服务器重启后状态自然归零，无需持久化。这种设计简化了实现。
spider_state = {
    'is_running': False,       # 爬虫是否正在运行
    'status': 'idle',          # 状态: idle(空闲), crawling(爬取中), analyzing(分析中), completed(完成), error(出错)
    'message': '',             # 当前状态描述信息
    'progress': 0              # 进度百分比 (0~100)
}

# 应用启动时自动创建数据库表（如果表不存在）
with app.app_context():
    db.create_all()


# ==============================================================================
# [路由-预测] 单条文本情感预测
# ==============================================================================
@app.route('/predict', methods=['POST'])
def predict():
    """
    功能：对单条文本进行情感分析，返回预测标签和置信度

    答辩要点：
        这是系统最基础的 API，演示了"前端 → 后端 → BERT模型 → 后端 → 前端"
        的完整数据链路。在实际应用中，可用于实时分析用户输入的文本情感。

    参数:
        POST body (JSON): {"text": "要分析的文本"}

    返回值:
        JSON: {"success": True, "text": "...", "sentiment": {"label": "...", "confidence": 0.95}}
    """
    try:
        data = request.get_json()

        if not data or 'text' not in data:
            return jsonify({
                'error': 'Missing required field: text'
            }), 400

        text = data['text']
        result = analyzer.predict(text)

        return jsonify({
            'success': True,
            'text': text,
            'sentiment': result
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-健康] 健康检查接口
# ==============================================================================
@app.route('/health', methods=['GET'])
def health():
    """
    功能：检查服务是否正常运行，返回模型加载状态

    答辩要点：
        健康检查接口是微服务架构的标准实践，运维系统通过定期请求此接口
        来判断服务是否存活。返回模型名称便于确认当前使用的模型版本。

    参数:
        无

    返回值:
        JSON: {"status": "healthy", "model": "模型路径"}
    """
    return jsonify({
        'status': 'healthy',
        'model': Config.BERT_MODEL_NAME
    }), 200


# ==============================================================================
# [路由-分析] 批量分析指定话题的评论
# ==============================================================================
@app.route('/analyze', methods=['POST'])
def analyze_comments():
    """
    功能：对指定话题的评论进行批量情感分析

    答辩要点：
        此接口实现了"懒分析"策略——评论先由爬虫存入数据库（无情感标签），
        调用此接口时才触发 BERT 推理。这种"存储-分析"分离的设计使得
        爬虫和模型互不阻塞，提高了系统的健壮性。

    参数:
        POST body (JSON): {"topic": "话题关键词", "limit": 10}

    返回值:
        JSON: {"success": True, "topic": "...", "count": N, "results": [...]}
    """
    try:
        data = request.get_json()

        if not data or 'topic' not in data:
            return jsonify({
                'error': 'Missing required field: topic'
            }), 400

        topic = data['topic']
        limit = data.get('limit', 10)

        comments = Comment.get_by_topic(topic, limit)
        results = []

        for comment in comments:
            if comment.sentiment_label is None:
                # 未分析的评论：调用 BERT 推理
                sentiment = analyzer.predict(comment.content)
                Comment.update_sentiment(
                    comment.id,
                    sentiment['label'],
                    sentiment[sentiment['label']]
                )
                results.append({
                    'id': comment.id,
                    'content': comment.content,
                    'sentiment': sentiment
                })
            else:
                # 已分析的评论：直接使用缓存结果
                results.append({
                    'id': comment.id,
                    'content': comment.content,
                    'sentiment': {
                        'label': comment.sentiment_label,
                        'confidence': comment.confidence
                    }
                })

        return jsonify({
            'success': True,
            'topic': topic,
            'count': len(results),
            'results': results
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-统计] 情感统计数据（饼图 + 折线图）
# ==============================================================================
@app.route('/api/sentiment_stats/<topic>', methods=['GET'])
def sentiment_stats(topic):
    """
    功能：获取指定话题的情感统计数据，供前端 ECharts 饼图和折线图使用

    答辩要点：
        此接口是"数据可视化"的核心数据源。它将数据库中的原始评论记录
        聚合为两种维度的统计：
        1. 饼图数据——情感分布（积极/中性/消极各多少条）
        2. 折线图数据——时序变化（按小时聚合平均情感值）
        这种"后端聚合、前端渲染"的架构，将计算压力放在服务器端，
        前端只负责展示，保证了页面的流畅度。

    参数:
        topic: URL路径参数，话题关键词

    返回值:
        JSON: {
            "success": True,
            "topic": "...",
            "total": 100,
            "pie_chart": {"data": [{"name": "积极", "value": 60}, ...]},
            "line_chart": {"data": [{"time": "2024-03-25 10:00", "value": 0.5}, ...]}
        }
    """
    try:
        comments = Comment.query.filter_by(topic_keyword=topic).all()

        if not comments:
            return jsonify({
                'success': False,
                'error': 'Topic not found'
            }), 404

        total = len(comments)
        positive_count = sum(1 for c in comments if c.sentiment_label == 'positive')
        neutral_count = sum(1 for c in comments if c.sentiment_label == 'neutral')
        negative_count = sum(1 for c in comments if c.sentiment_label == 'negative')

        # 饼图数据：情感分布
        pie_data = [
            {'name': '积极', 'value': positive_count},
            {'name': '中性', 'value': neutral_count},
            {'name': '消极', 'value': negative_count}
        ]

        # 折线图数据：将情感标签映射为数值，便于计算平均值
        # positive=1, neutral=0, negative=-1
        # 答辩要点：这种映射方式使得平均值的含义直观——
        # 大于0表示整体偏积极，小于0表示偏消极，接近0表示中性
        sentiment_map = {'positive': 1, 'neutral': 0, 'negative': -1}

        # 按小时分组，计算每小时的平均情感值
        time_groups = {}
        for comment in comments:
            if comment.sentiment_label and comment.sentiment_label in sentiment_map:
                # 将时间截断到小时级别，如 "2024-03-25 10:00"
                hour_key = comment.create_time.strftime('%Y-%m-%d %H:00')
                if hour_key not in time_groups:
                    time_groups[hour_key] = []
                time_groups[hour_key].append(sentiment_map[comment.sentiment_label])

        # 按时间排序，计算每小时的平均情感值
        sorted_times = sorted(time_groups.keys())
        line_data = []
        for time_key in sorted_times:
            values = time_groups[time_key]
            avg_sentiment = sum(values) / len(values) if values else 0
            line_data.append({
                'time': time_key,
                'value': round(avg_sentiment, 2)
            })

        return jsonify({
            'success': True,
            'topic': topic,
            'total': total,
            'pie_chart': {
                'data': pie_data
            },
            'line_chart': {
                'data': line_data
            }
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-词云] 词云数据接口
# ==============================================================================
@app.route('/api/wordcloud/<topic>', methods=['GET'])
def wordcloud_stats(topic):
    """
    功能：获取指定话题的高频词汇数据，供前端 ECharts 词云图使用

    答辩要点：
        使用 jieba 库的 TextRank 算法提取关键词，而非简单的词频统计。
        TextRank 基于图排序算法（类似 PageRank），能识别出文本中
        "重要性高"的词汇，而不仅仅是出现次数多的停用词。
        allowPOS 参数限制只提取名词(n)、人名(nr)、地名(ns)、
        动名词(vn)、动词(v)，过滤掉无意义的虚词。

    参数:
        topic: URL路径参数，话题关键词

    返回值:
        JSON: {"success": True, "topic": "...", "wordcloud": {"data": [{"name": "词汇", "value": 权重}, ...]}}
    """
    try:
        comments = Comment.query.filter_by(topic_keyword=topic).all()

        if not comments:
            return jsonify({
                'success': False,
                'error': 'Topic not found'
            }), 404

        # 将所有评论拼接为一个长字符串
        content = " ".join([comment.content for comment in comments])

        # [分词] 使用 TextRank 算法提取关键词
        # topK=100: 最多提取100个关键词
        # withWeight=True: 返回词汇及其权重
        # allowPOS: 只提取名词、人名、地名、动名词、动词
        tags = jieba.analyse.textrank(
            content,
            topK=100,
            withWeight=True,
            allowPOS=('n', 'nr', 'ns', 'vn', 'v')
        )

        # 将权重乘以1000取整，作为词云的"字体大小"参考值
        top_words = [{'name': word, 'value': int(weight * 1000)} for word, weight in tags]

        return jsonify({
            'success': True,
            'topic': topic,
            'wordcloud': {
                'data': top_words
            }
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-页面] 首页和详情页
# ==============================================================================
@app.route('/')
def index():
    """
    功能：渲染首页（话题列表页）

    参数:
        无

    返回值:
        HTML 页面
    """
    return render_template('index.html')


@app.route('/detail/<topic>')
def detail(topic):
    """
    功能：渲染话题详情页（可视化图表页）

    参数:
        topic: URL路径参数，话题关键词

    返回值:
        HTML 页面（携带 topic 变量供前端 JS 使用）
    """
    return render_template('detail.html', topic=topic)


# ==============================================================================
# [路由-导出] 导出分析结果（CSV / Excel）
# ==============================================================================
@app.route('/export/<topic>/<format>', methods=['GET'])
def export_data(topic, format):
    """
    功能：导出指定话题的情感分析结果为 CSV 或 Excel 文件

    答辩要点：
        1. CSV 导出时添加 UTF-8 BOM (\\xef\\xbb\\xbf)，解决 Excel
           打开中文乱码问题——Excel 默认用系统编码（GBK）打开 CSV，
           BOM 头强制 Excel 识别为 UTF-8。
        2. 文件名使用 RFC 5987 标准编码（filename*=utf-8''...），
           解决中文文件名在 HTTP 头中导致的 UnicodeEncodeError。
        3. Excel 导出使用 openpyxl 引擎，支持单元格样式设置
           （表头蓝底白字、情感标签条件着色）。

    参数:
        topic: URL路径参数，话题关键词
        format: URL路径参数，导出格式 ('csv' 或 'excel')

    返回值:
        文件下载响应（CSV 或 Excel）
    """
    try:
        comments = Comment.query.filter_by(topic_keyword=topic).all()

        if not comments:
            return jsonify({
                'success': False,
                'error': 'Topic not found'
            }), 404

        # 构建导出数据
        data = []
        for comment in comments:
            data.append({
                'ID': comment.id,
                '话题': comment.topic_keyword,
                '评论内容': comment.content,
                '发布时间': comment.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                '情感标签': comment.sentiment_label if comment.sentiment_label else '未分析',
                '置信度': round(comment.confidence, 4) if comment.confidence else 0,
                '手动标注': comment.manual_label if comment.manual_label else '无'
            })

        df = pd.DataFrame(data)

        # 将英文标签映射为中文，方便阅读
        sentiment_map = {'positive': '积极', 'neutral': '中性', 'negative': '消极'}
        df['情感标签'] = df['情感标签'].map(lambda x: sentiment_map.get(x, x))
        df['手动标注'] = df['手动标注'].map(lambda x: sentiment_map.get(x, x))

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"weibo_sentiment_{topic}_{timestamp}"

        if format.lower() == 'csv':
            output = io.BytesIO()
            # 写入 UTF-8 BOM，解决 Excel 中文乱码
            output.write(b'\xef\xbb\xbf')
            csv_content = df.to_csv(index=False, encoding='utf-8')
            output.write(csv_content.encode('utf-8'))
            output.seek(0)

            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            # RFC 5987 标准编码中文文件名
            encoded_filename = quote(f"{filename}.csv")
            response.headers['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
            return response

        elif format.lower() == 'excel':
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='情感分析数据')

                worksheet = writer.sheets['情感分析数据']

                # 表头样式：蓝底白字加粗
                header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)

                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 自动调整列宽
                    column_letter = get_column_letter(col_num)
                    max_length = max(
                        df[column].astype(str).apply(len).max(),
                        len(str(column))
                    )
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # 数据行样式：情感标签条件着色
                for row_num, row_data in enumerate(df.itertuples(index=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                        # 第5列是情感标签，根据标签着色
                        if col_num == 5:
                            if value == '积极':
                                cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                            elif value == '消极':
                                cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                            elif value == '中性':
                                cell.fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')

            output.seek(0)

            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            encoded_filename = quote(f"{filename}.xlsx")
            response.headers['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
            return response

        else:
            return jsonify({
                'success': False,
                'error': 'Unsupported format. Use "excel" or "csv"'
            }), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-摘要] 导出统计摘要
# ==============================================================================
@app.route('/export_summary/<topic>', methods=['GET'])
def export_summary(topic):
    """
    功能：导出指定话题的统计摘要为 Excel 文件

    答辩要点：
        统计摘要将原始评论聚合为宏观指标（总数、各情感占比、平均置信度），
        便于快速了解话题的整体情感倾向。这是"数据分析"层面的输出，
        与逐条导出形成互补。

    参数:
        topic: URL路径参数，话题关键词

    返回值:
        Excel 文件下载响应
    """
    try:
        comments = Comment.query.filter_by(topic_keyword=topic).all()

        if not comments:
            return jsonify({
                'success': False,
                'error': 'Topic not found'
            }), 404

        total = len(comments)
        positive_count = sum(1 for c in comments if c.sentiment_label == 'positive')
        neutral_count = sum(1 for c in comments if c.sentiment_label == 'neutral')
        negative_count = sum(1 for c in comments if c.sentiment_label == 'negative')

        avg_confidence = sum(c.confidence for c in comments if c.confidence) / total if total > 0 else 0

        summary_data = {
            '话题': [topic],
            '总评论数': [total],
            '积极评论': [positive_count],
            '积极占比': [f"{positive_count/total*100:.2f}%"] if total > 0 else ['0%'],
            '中性评论': [neutral_count],
            '中性占比': [f"{neutral_count/total*100:.2f}%"] if total > 0 else ['0%'],
            '消极评论': [negative_count],
            '消极占比': [f"{negative_count/total*100:.2f}%"] if total > 0 else ['0%'],
            '平均置信度': [f"{avg_confidence:.4f}"],
            '导出时间': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        }

        df = pd.DataFrame(summary_data)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"weibo_summary_{topic}_{timestamp}"

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='统计摘要')

            worksheet = writer.sheets['统计摘要']

            # 表头样式：紫底白字加粗
            header_fill = PatternFill(start_color='764ba2', end_color='764ba2', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=14)

            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

                column_letter = get_column_letter(col_num)
                max_length = max(
                    df[column].astype(str).apply(len).max(),
                    len(str(column))
                )
                adjusted_width = min(max_length + 4, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=12)

        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        encoded_filename = quote(f"{filename}.xlsx")
        response.headers['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
        return response

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-爬虫] 启动爬虫（SSE 流式响应）
# ==============================================================================
@app.route('/api/run_spider', methods=['POST'])
def run_spider():
    """
    功能：启动微博爬虫并自动进行情感分析，通过 SSE 实时推送进度

    答辩要点：
        此接口是系统最复杂的 API，采用了三项关键技术：
        1. SSE (Server-Sent Events)：实现服务器到客户端的单向实时推送，
           比 WebSocket 更轻量，适合"服务器推送进度"这种场景。
        2. 多线程：爬虫在子线程中运行，主线程通过队列接收进度信息，
           避免爬虫阻塞 Flask 的请求处理线程。
        3. 流式响应：Response + stream_with_context + generator，
           让 HTTP 连接保持打开状态，持续发送进度数据。

    参数:
        POST body (JSON):
            mode: 'hot' (热搜榜单) 或 'search' (自定义搜索)
            keyword: 搜索关键词 (仅在 search 模式下有效)
            topic_count: 爬取的话题数量 (默认: 5, 范围: 1-20)
            count_per_topic: 每个话题爬取的评论数量 (默认: 150, 范围: 10-300)
            clear_history: 是否在爬取前清空所有历史数据 (默认: False)

    返回值:
        SSE 流: data: {"progress": 50.0, "message": "...", "status": "processing"}\n\n
    """
    try:
        data = request.get_json()
        mode = data.get('mode', 'hot')
        keyword = data.get('keyword', '').strip()
        topic_count = data.get('topic_count', 5)
        count_per_topic = data.get('count_per_topic', 150)
        clear_history = data.get('clear_history', False)

        # 参数验证
        if mode not in ['hot', 'search']:
            return jsonify({
                'success': False,
                'error': '模式必须是 hot 或 search'
            }), 400

        if mode == 'search' and not keyword:
            return jsonify({
                'success': False,
                'error': '搜索模式下必须提供关键词'
            }), 400

        if topic_count < 1 or topic_count > 20:
            return jsonify({
                'success': False,
                'error': '话题数量必须在 1-20 之间'
            }), 400

        if count_per_topic < 10 or count_per_topic > 300:
            return jsonify({
                'success': False,
                'error': '评论数量必须在 10-300 之间'
            }), 400

        # 创建队列用于传递进度信息（线程间通信）
        progress_queue = queue.Queue()

        def send_progress(progress, message, status='processing', error=None):
            """
            功能：发送进度更新到队列

            答辩要点：此函数是子线程和主线程之间的"通信桥梁"。
            子线程调用此函数将进度信息放入队列，主线程的 generate()
            函数从队列取出并通过 SSE 推送给前端。

            参数:
                progress: float —— 进度百分比 (0~100)
                message: str —— 进度描述
                status: str —— 状态标识
                error: str or None —— 错误信息
            """
            import json

            # 同步更新全局状态字典
            spider_state['progress'] = round(progress, 1)
            spider_state['message'] = message
            spider_state['status'] = status
            spider_state['is_running'] = (status not in ['completed', 'error'])

            data = {
                'progress': round(progress, 1),
                'message': message,
                'status': status
            }
            if error:
                data['error'] = error
            # SSE 格式: "data: JSON内容\n\n"
            progress_queue.put(f"data: {json.dumps(data, ensure_ascii=False)}\n\n")

        def spider_task():
            """
            功能：爬虫任务（在子线程中运行）

            答辩要点：
                整个爬虫流程分为4个阶段，进度分配如下：
                - 阶段1 (0-10%): 初始化浏览器
                - 阶段2 (10-60%): 爬取评论
                - 阶段3 (60-90%): BERT 情感分析
                - 阶段4 (90-100%): 完成
                这种分阶段进度设计让用户能清楚知道当前执行到哪一步。

            参数:
                无（通过闭包访问外部变量）

            返回值:
                无（通过队列传递结果）
            """
            try:
                from weibo_spider import WeiboSpider
                from model_inference import SentimentAnalyzer

                # 如果需要清空历史数据
                if clear_history:
                    send_progress(0, "正在清空所有历史数据...")
                    try:
                        with app.app_context():
                            deleted_count = Comment.query.delete()
                            db.session.commit()
                            send_progress(0, f"已清空 {deleted_count} 条历史数据")
                    except Exception as e:
                        send_progress(0, f"清空历史数据失败: {str(e)}", status='error', error=str(e))
                        progress_queue.put(None)
                        return

                # 创建爬虫实例（headless=False 显示浏览器窗口）
                spider = WeiboSpider(headless=False)

                # 根据模式选择爬取方式
                if mode == 'search':
                    send_progress(0, f"开始搜索关键词：{keyword}")
                    total_saved = spider.crawl_single_topic(
                        keyword,
                        count_per_topic,
                        progress_callback=send_progress
                    )
                else:
                    send_progress(0, "开始获取微博热搜榜...")
                    total_saved = spider.crawl_hot_topics(
                        count_per_topic,
                        topic_count,
                        progress_callback=send_progress
                    )

                # 阶段3: BERT分析 (60-90%)
                send_progress(60, f"爬取完成！共获取 {total_saved} 条评论，开始情感分析...")

                sentiment_analyzer = SentimentAnalyzer()

                # 运行情感分析（带进度回调）
                analyzed_count = sentiment_analyzer.batch_analyze_database(
                    batch_size=32,
                    progress_callback=send_progress
                )

                # 阶段4: 完成 (100%)
                send_progress(100, f"分析完成！共处理 {total_saved} 条评论，其中 {analyzed_count} 条已完成情感标注。", status='completed')

                # 发送结束信号（None 表示流结束）
                progress_queue.put(None)

            except Exception as e:
                import traceback
                error_msg = f"爬虫任务出错: {str(e)}"
                print(f"错误详情: {traceback.format_exc()}")
                send_progress(0, error_msg, status='error', error=str(e))
                progress_queue.put(None)

        def generate():
            """
            功能：生成 SSE 流式响应的生成器函数

            答辩要点：
                这是 Flask SSE 的标准实现模式：
                1. 启动子线程执行耗时任务
                2. 主线程通过 yield 持续发送数据
                3. 当收到 None 信号时结束流
                关键技术：stream_with_context() 保持 Flask 请求上下文，
                否则在子线程中无法访问数据库等 Flask 资源。

            参数:
                无

            返回值:
                generator —— 产出 SSE 格式的字符串
            """
            # 在子线程中运行爬虫任务
            thread = threading.Thread(target=spider_task)
            thread.start()

            # 发送初始消息
            send_progress(0, "正在启动爬虫任务...")

            # 从队列读取进度信息并返回
            while True:
                try:
                    # 设置超时以避免永久阻塞
                    progress_data = progress_queue.get(timeout=1)

                    if progress_data is None:
                        # 收到结束信号，退出循环
                        break

                    yield progress_data

                except queue.Empty:
                    # 队列为空，继续等待
                    continue

            # 等待子线程完成
            thread.join()

        # 返回 SSE 流式响应
        # 答辩要点：mimetype='text/event-stream' 是 SSE 的标准 MIME 类型，
        # 浏览器据此知道这是一个持续推送的事件流。
        # X-Accel-Buffering: no 禁止 Nginx 缓冲，确保数据实时到达前端。
        return Response(
            stream_with_context(generate()),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no'
            }
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-状态] 获取爬虫运行状态
# ==============================================================================
@app.route('/api/spider_status', methods=['GET'])
def spider_status():
    """
    功能：获取爬虫当前运行状态和数据库统计信息

    答辩要点：
        此接口供前端轮询，实时展示爬虫状态和数据库中的评论数量。
        返回已分析/未分析的评论数，便于用户判断是否需要重新分析。

    参数:
        无

    返回值:
        JSON: {"success": True, "spider_state": {...}, "total_comments": N, "analyzed_comments": M, "unanalyzed_comments": K}
    """
    try:
        with app.app_context():
            total_comments = Comment.query.count()
            analyzed_comments = Comment.query.filter(Comment.sentiment_label.isnot(None)).count()

            return jsonify({
                'success': True,
                'spider_state': spider_state,
                'total_comments': total_comments,
                'analyzed_comments': analyzed_comments,
                'unanalyzed_comments': total_comments - analyzed_comments
            }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# [路由-话题] 获取已有话题列表
# ==============================================================================
@app.route('/api/hot_topics', methods=['GET'])
def get_hot_topics():
    """
    功能：获取数据库中所有话题，按最新创建时间倒序排列

    答辩要点：
        使用 SQLAlchemy 的 group_by + max 聚合查询，按话题分组并
        取每组的最大创建时间，然后按时间倒序排列。这样用户看到的
        话题列表总是"最近活跃"的排在前面，符合信息时效性需求。

    参数:
        无

    返回值:
        JSON: {"success": True, "topics": ["话题1", "话题2", ...]}
    """
    try:
        with app.app_context():
            from sqlalchemy import func

            # 查询每个话题的最新创建时间，并按时间倒序排列
            topics_with_time = db.session.query(
                Comment.topic_keyword,
                func.max(Comment.create_time).label('latest_time')
            ).group_by(Comment.topic_keyword).order_by(
                func.max(Comment.create_time).desc()
            ).all()

            topic_list = [topic[0] for topic in topics_with_time]

            return jsonify({
                'success': True,
                'topics': topic_list
            }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# 应用启动入口
# ==============================================================================
if __name__ == '__main__':
    print(f"Starting Flask server on {Config.FLASK_HOST}:{Config.FLASK_PORT}")
    app.run(
        host=Config.FLASK_HOST,
        port=Config.FLASK_PORT,
        debug=Config.FLASK_DEBUG
    )
